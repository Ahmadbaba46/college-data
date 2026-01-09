from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Count

from courses.models import Course, CourseOffering
from students.models import Level
from core.models import Department
from users.decorators import admin_or_data_entry_required


@login_required
def course_list(request):
    """
    List all courses with search and filters.
    All authenticated users can view.
    """
    courses = Course.objects.select_related('department').annotate(
        offerings_count=Count('offerings')
    ).order_by('code')
    
    # Search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        courses = courses.filter(
            Q(code__icontains=search_query) |
            Q(title__icontains=search_query)
        )
    
    # Filter by level
    level_id = request.GET.get('level')
    if level_id:
        courses = courses.filter(levels__id=level_id)
    
    # Filter by semester
    semester = request.GET.get('semester')
    if semester:
        courses = courses.filter(default_semester=semester)
    
    # Filter by department
    department_id = request.GET.get('department')
    if department_id:
        courses = courses.filter(department_id=department_id)
    
    # Pagination
    paginator = Paginator(courses, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_obj': page_obj,
        'levels': Level.objects.all(),
        'departments': departments,
        'search_query': search_query,
        'total_count': courses.count(),
    }
    
    return render(request, 'courses/list.html', context)


@login_required
def course_detail(request, course_id):
    """
    View course details with offerings.
    All authenticated users can view.
    """
    from academics.models import CurriculumCourse, Program
    
    course = get_object_or_404(
        Course.objects.prefetch_related('levels', 'sessions'),
        id=course_id
    )
    
    offerings = CourseOffering.objects.filter(
        course=course
    ).select_related('session', 'level').order_by('-session__name')
    
    # Get programs that include this course in their curriculum
    curriculum_entries = CurriculumCourse.objects.filter(
        course=course
    ).select_related('program', 'level').order_by('program__name', 'level__name', 'semester')
    
    # Get all programs and levels for quick-add form
    programs = Program.objects.all().order_by('name')
    levels = Level.objects.all()
    
    context = {
        'course': course,
        'offerings': offerings,
        'curriculum_entries': curriculum_entries,
        'programs': programs,
        'levels': levels,
    }
    
    return render(request, 'courses/detail.html', context)


@admin_or_data_entry_required
def course_create(request):
    """
    Create new course.
    Only Admin and DataEntry can access.
    """
    if request.method == 'POST':
        try:
            department_id = request.POST.get('department')
            department = None
            if department_id:
                department = Department.objects.filter(id=department_id).first()
            
            course = Course.objects.create(
                code=request.POST.get('code'),
                title=request.POST.get('title'),
                units=request.POST.get('units', 3),
                department=department,
                default_semester=request.POST.get('default_semester') or None,
            )
            
            messages.success(request, f'Course {course.code} created successfully!')
            return redirect('courses:detail', course_id=course.id)
        
        except Exception as e:
            messages.error(request, f'Error creating course: {str(e)}')
    
    context = {
        'levels': Level.objects.all(),
        'departments': Department.objects.filter(is_active=True).order_by('name'),
    }
    
    return render(request, 'courses/create.html', context)


@admin_or_data_entry_required
def course_edit(request, course_id):
    """
    Edit existing course.
    Only Admin and DataEntry can access.
    """
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        try:
            department_id = request.POST.get('department')
            department = None
            if department_id:
                department = Department.objects.filter(id=department_id).first()
            
            course.code = request.POST.get('code')
            course.title = request.POST.get('title')
            course.units = request.POST.get('units', 3)
            course.department = department
            course.default_semester = request.POST.get('default_semester') or None
            course.save()
            
            messages.success(request, f'Course {course.code} updated successfully!')
            return redirect('courses:detail', course_id=course.id)
        
        except Exception as e:
            messages.error(request, f'Error updating course: {str(e)}')
    
    context = {
        'course': course,
        'levels': Level.objects.all(),
        'departments': Department.objects.filter(is_active=True).order_by('name'),
    }
    
    return render(request, 'courses/edit.html', context)


@admin_or_data_entry_required
def course_delete(request, course_id):
    """
    Delete course.
    Only Admin and DataEntry can access.
    """
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        course_code = course.code
        course.delete()
        messages.success(request, f'Course {course_code} deleted successfully!')
        return redirect('courses:list')
    
    return render(request, 'courses/delete_confirm.html', {'course': course})


@admin_or_data_entry_required
def course_import(request):
    """Bulk course import from CSV/Excel/JSON"""
    import csv
    import json
    from io import TextIOWrapper
    
    if request.method == 'POST':
        uploaded_file = request.FILES.get('import_file')
        
        if not uploaded_file:
            messages.error(request, 'Please select a file to import.')
            return redirect('courses:import')
        
        file_ext = uploaded_file.name.split('.')[-1].lower()
        
        # Parse file based on extension
        try:
            imported = 0
            errors = []
            
            if file_ext == 'csv':
                # Handle CSV
                file_data = TextIOWrapper(uploaded_file.file, encoding='utf-8')
                reader = csv.DictReader(file_data)
                
                for idx, row in enumerate(reader, start=2):
                    result = _import_course_row(row, idx)
                    if result['success']:
                        imported += 1
                    else:
                        errors.append(result['error'])
                        
            elif file_ext in ['xlsx', 'xls']:
                # Handle Excel
                from openpyxl import load_workbook
                workbook = load_workbook(uploaded_file)
                sheet = workbook.active
                header = [cell.value for cell in sheet[1]]
                
                for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    row_data = {header[i]: cell.value for i, cell in enumerate(row) if i < len(header)}
                    result = _import_course_row(row_data, idx)
                    if result['success']:
                        imported += 1
                    else:
                        errors.append(result['error'])
                        
            elif file_ext == 'json':
                # Handle JSON
                data = json.load(uploaded_file)
                for idx, row in enumerate(data, start=1):
                    result = _import_course_row(row, idx)
                    if result['success']:
                        imported += 1
                    else:
                        errors.append(result['error'])
            else:
                messages.error(request, 'Unsupported file format. Please use CSV, Excel, or JSON.')
                return redirect('courses:import')
            
            # Show results
            if imported > 0:
                messages.success(request, f'Successfully imported {imported} course(s).')
            
            if errors:
                error_msg = f'{len(errors)} error(s) occurred during import. '
                messages.warning(request, error_msg + ' | '.join(errors[:5]))
                if len(errors) > 5:
                    messages.info(request, f'... and {len(errors) - 5} more errors.')
            
            return redirect('courses:list')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('courses:import')
    
    # GET request - show form with sample templates
    return render(request, 'courses/import.html')


def _import_course_row(row, row_num):
    """Helper function to import a single course row"""
    from academics.models import Program, CurriculumCourse
    
    try:
        # Required fields
        required_fields = ['code', 'title']
        for field in required_fields:
            if not row.get(field):
                return {'success': False, 'error': f'Row {row_num}: Missing required field "{field}"'}
        
        # Create or update course
        course, created = Course.objects.update_or_create(
            code=row['code'],
            defaults={
                'title': row['title'],
                'units': int(row.get('units', 3)),
                'default_semester': row.get('default_semester', '').upper() if row.get('default_semester') else None,
            }
        )
        
        # Optional: Add to program curriculum
        program_code = row.get('program_code')
        if program_code:
            try:
                program = Program.objects.get(code=program_code)
                level_name = row.get('level', '100 Level')
                level = Level.objects.get(name=level_name)
                semester = row.get('semester', 'FIRST').upper()
                is_compulsory = str(row.get('is_compulsory', 'yes')).lower() in ['yes', 'true', '1', 'compulsory']
                
                CurriculumCourse.objects.get_or_create(
                    program=program,
                    course=course,
                    level=level,
                    semester=semester,
                    defaults={'is_compulsory': is_compulsory}
                )
            except (Program.DoesNotExist, Level.DoesNotExist):
                # Skip curriculum linking if program or level not found
                pass
        
        return {'success': True, 'course': course, 'created': created}
        
    except Exception as e:
        return {'success': False, 'error': f'Row {row_num}: {str(e)}'}


@admin_or_data_entry_required
def offering_create(request, course_id):
    """Create a new course offering"""
    from students.models import Session
    from teachers.models import Teacher
    
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        try:
            teacher_id = request.POST.get('teacher')
            offering = CourseOffering.objects.create(
                course=course,
                session_id=request.POST.get('session'),
                semester=request.POST.get('semester'),
                level_id=request.POST.get('level') or None,
                teacher_id=teacher_id if teacher_id else None,
                capacity=request.POST.get('capacity') or None,
                is_active=request.POST.get('is_active') == 'on',
            )
            
            messages.success(request, f'Course offering created successfully!')
            return redirect('courses:detail', course_id=course.id)
        
        except Exception as e:
            messages.error(request, f'Error creating offering: {str(e)}')
    
    context = {
        'course': course,
        'sessions': Session.objects.all().order_by('-name'),
        'levels': Level.objects.all(),
        'teachers': Teacher.objects.filter(is_active=True).order_by('last_name', 'first_name'),
    }
    
    return render(request, 'courses/offering_create.html', context)


@admin_or_data_entry_required
def offering_edit(request, offering_id):
    """Edit an existing course offering"""
    from students.models import Session
    from teachers.models import Teacher
    
    offering = get_object_or_404(CourseOffering, id=offering_id)
    
    if request.method == 'POST':
        try:
            teacher_id = request.POST.get('teacher')
            offering.session_id = request.POST.get('session')
            offering.semester = request.POST.get('semester')
            offering.level_id = request.POST.get('level') or None
            offering.teacher_id = teacher_id if teacher_id else None
            offering.capacity = request.POST.get('capacity') or None
            offering.is_active = request.POST.get('is_active') == 'on'
            offering.save()
            
            messages.success(request, f'Course offering updated successfully!')
            return redirect('courses:detail', course_id=offering.course.id)
        
        except Exception as e:
            messages.error(request, f'Error updating offering: {str(e)}')
    
    context = {
        'offering': offering,
        'course': offering.course,
        'sessions': Session.objects.all().order_by('-name'),
        'levels': Level.objects.all(),
        'teachers': Teacher.objects.filter(is_active=True).order_by('last_name', 'first_name'),
    }
    
    return render(request, 'courses/offering_edit.html', context)


@admin_or_data_entry_required
def offering_delete(request, offering_id):
    """Delete a course offering"""
    offering = get_object_or_404(CourseOffering, id=offering_id)
    course_id = offering.course.id
    
    if request.method == 'POST':
        offering.delete()
        messages.success(request, f'Course offering deleted successfully!')
        return redirect('courses:detail', course_id=course_id)
    
    return render(request, 'courses/offering_delete_confirm.html', {'offering': offering})


@admin_or_data_entry_required
def manage_curriculum(request, course_id):
    """Manage course curriculum - add/remove from programs"""
    from academics.models import Program, CurriculumCourse
    from students.models import Session
    
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            try:
                CurriculumCourse.objects.create(
                    course=course,
                    program_id=request.POST.get('program'),
                    level_id=request.POST.get('level'),
                    semester=request.POST.get('semester'),
                    is_compulsory=request.POST.get('is_compulsory') == 'on',
                )
                messages.success(request, f'Course added to program curriculum!')
            except Exception as e:
                messages.error(request, f'Error adding to curriculum: {str(e)}')
        
        elif action == 'remove':
            try:
                entry_id = request.POST.get('entry_id')
                CurriculumCourse.objects.filter(id=entry_id).delete()
                messages.success(request, f'Course removed from program curriculum!')
            except Exception as e:
                messages.error(request, f'Error removing from curriculum: {str(e)}')
        
        return redirect('courses:manage_curriculum', course_id=course.id)
    
    # GET request
    curriculum_entries = CurriculumCourse.objects.filter(
        course=course
    ).select_related('program', 'level').order_by('program__name', 'level__name', 'semester')
    
    programs = Program.objects.all().order_by('name')
    levels = Level.objects.all()
    
    context = {
        'course': course,
        'curriculum_entries': curriculum_entries,
        'programs': programs,
        'levels': levels,
    }
    
    return render(request, 'courses/manage_curriculum.html', context)


@admin_or_data_entry_required
def quick_add_to_program(request, course_id):
    """Quick add course to program via HTMX"""
    from academics.models import Program, CurriculumCourse
    from django.http import HttpResponse
    
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        try:
            CurriculumCourse.objects.create(
                course=course,
                program_id=request.POST.get('program'),
                level_id=request.POST.get('level'),
                semester=request.POST.get('semester'),
                is_compulsory=request.POST.get('is_compulsory') == 'on',
            )
            # Return updated list
            curriculum_entries = CurriculumCourse.objects.filter(
                course=course
            ).select_related('program', 'level').order_by('program__name', 'level__name', 'semester')
            
            return render(request, 'courses/partials/curriculum_list.html', {
                'curriculum_entries': curriculum_entries,
                'course': course,
            })
        except Exception as e:
            return HttpResponse(f'<div class="text-red-600 text-sm">Error: {str(e)}</div>')
    
    return HttpResponse('')
