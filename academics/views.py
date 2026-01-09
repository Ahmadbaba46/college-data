from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count

from academics.models import Program, CurriculumCourse, Prerequisite
from students.models import Student, Level
from courses.models import Course
from core.models import Department
from users.decorators import admin_required


@login_required
def program_list(request):
    """
    List all programs with student counts.
    All authenticated users can view.
    """
    programs = Program.objects.select_related('department').annotate(
        student_count=Count('students')
    ).order_by('code')
    
    # Filter by department
    department_id = request.GET.get('department')
    if department_id:
        programs = programs.filter(department_id=department_id)
    
    # Search by code or name
    search_query = request.GET.get('search', '').strip()
    if search_query:
        from django.db.models import Q
        programs = programs.filter(
            Q(code__icontains=search_query) |
            Q(name__icontains=search_query)
        )
    
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    context = {
        'programs': programs,
        'departments': departments,
        'search_query': search_query,
    }
    
    return render(request, 'academics/program_list.html', context)


@login_required
def program_detail(request, program_id):
    """
    View program details with curriculum.
    All authenticated users can view.
    """
    program = get_object_or_404(Program, id=program_id)
    
    # Get curriculum courses
    curriculum_courses = CurriculumCourse.objects.filter(
        program=program
    ).select_related('course', 'level').order_by('level__name', 'course__code')
    
    # Get students in this program
    students = Student.objects.filter(program=program).select_related('current_level')[:10]
    student_count = Student.objects.filter(program=program).count()
    
    context = {
        'program': program,
        'curriculum_courses': curriculum_courses,
        'students': students,
        'student_count': student_count,
    }
    
    return render(request, 'academics/program_detail.html', context)


@admin_required
def program_create(request):
    """
    Create new program.
    Only Admin can access.
    """
    if request.method == 'POST':
        try:
            department_id = request.POST.get('department')
            department = None
            if department_id:
                department = Department.objects.filter(id=department_id).first()
            
            program = Program.objects.create(
                code=request.POST.get('code'),
                name=request.POST.get('name'),
                department=department,
                min_units_to_graduate=request.POST.get('min_units_to_graduate', 0),
                classification_scheme=request.POST.get('classification_scheme', Program.CLASS_SCHEME_BSC),
            )
            
            messages.success(request, f'Program {program.code} created successfully!')
            return redirect('academics:program_detail', program_id=program.id)
        
        except Exception as e:
            messages.error(request, f'Error creating program: {str(e)}')
    
    departments = Department.objects.filter(is_active=True).order_by('name')
    return render(request, 'academics/program_create.html', {'departments': departments})


@admin_required
def program_edit(request, program_id):
    """
    Edit existing program.
    Only Admin can access.
    """
    program = get_object_or_404(Program, id=program_id)
    
    if request.method == 'POST':
        try:
            department_id = request.POST.get('department')
            department = None
            if department_id:
                department = Department.objects.filter(id=department_id).first()
            
            program.code = request.POST.get('code')
            program.name = request.POST.get('name')
            program.department = department
            program.min_units_to_graduate = request.POST.get('min_units_to_graduate', 0)
            program.classification_scheme = request.POST.get('classification_scheme', Program.CLASS_SCHEME_BSC)
            program.save()
            
            messages.success(request, f'Program {program.code} updated successfully!')
            return redirect('academics:program_detail', program_id=program.id)
        
        except Exception as e:
            messages.error(request, f'Error updating program: {str(e)}')
    
    departments = Department.objects.filter(is_active=True).order_by('name')
    context = {'program': program, 'departments': departments}
    return render(request, 'academics/program_edit.html', context)


@admin_required
def curriculum_add(request, program_id):
    """Add a course to program curriculum"""
    program = get_object_or_404(Program, id=program_id)
    
    if request.method == 'POST':
        # Check if bulk operation
        course_ids = request.POST.getlist('course_ids')
        
        if course_ids:
            # Bulk add
            level_id = request.POST.get('level_id')
            semester = request.POST.get('semester')
            is_compulsory = request.POST.get('is_compulsory') == 'on'
            
            try:
                level = Level.objects.get(id=level_id)
                added = 0
                skipped = 0
                
                for course_id in course_ids:
                    try:
                        course = Course.objects.get(id=course_id)
                        _, created = CurriculumCourse.objects.get_or_create(
                            program=program,
                            course=course,
                            level=level,
                            semester=semester,
                            defaults={'is_compulsory': is_compulsory}
                        )
                        if created:
                            added += 1
                        else:
                            skipped += 1
                    except Course.DoesNotExist:
                        continue
                
                messages.success(request, f'Added {added} course(s) to curriculum. {skipped} already existed.')
                return redirect('academics:program_detail', program_id=program.id)
                
            except Level.DoesNotExist:
                messages.error(request, 'Invalid level selected.')
                return redirect('academics:program_detail', program_id=program.id)
        else:
            # Single add
            try:
                course_id = request.POST.get('course_id')
                level_id = request.POST.get('level_id')
                semester = request.POST.get('semester')
                is_compulsory = request.POST.get('is_compulsory') == 'on'
                
                course = Course.objects.get(id=course_id)
                level = Level.objects.get(id=level_id)
                
                curriculum_course, created = CurriculumCourse.objects.get_or_create(
                    program=program,
                    course=course,
                    level=level,
                    semester=semester,
                    defaults={'is_compulsory': is_compulsory}
                )
                
                if created:
                    messages.success(request, f'Added {course.code} to {program.code} curriculum.')
                else:
                    curriculum_course.is_compulsory = is_compulsory
                    curriculum_course.save()
                    messages.info(request, f'Updated {course.code} in curriculum.')
                
                return redirect('academics:program_detail', program_id=program.id)
                
            except (Course.DoesNotExist, Level.DoesNotExist) as e:
                messages.error(request, f'Invalid course or level selected.')
                return redirect('academics:program_detail', program_id=program.id)
    
    # GET - show form
    courses = Course.objects.all().order_by('code')
    levels = Level.objects.all().order_by('name')
    
    ctx = {
        'program': program,
        'courses': courses,
        'levels': levels,
    }
    return render(request, 'academics/curriculum_add.html', ctx)


@admin_required
def curriculum_remove(request, program_id, curriculum_id):
    """Remove a course from program curriculum"""
    program = get_object_or_404(Program, id=program_id)
    curriculum_course = get_object_or_404(CurriculumCourse, id=curriculum_id, program=program)
    
    if request.method == 'POST':
        course_code = curriculum_course.course.code
        curriculum_course.delete()
        messages.success(request, f'Removed {course_code} from {program.code} curriculum.')
        return redirect('academics:program_detail', program_id=program.id)
    
    return redirect('academics:program_detail', program_id=program.id)


@admin_required
def prerequisites_manage(request, program_id):
    """View and manage prerequisites for a program"""
    program = get_object_or_404(Program, id=program_id)
    prerequisites = Prerequisite.objects.filter(program=program).select_related(
        'course', 'prerequisite_course'
    ).order_by('course__code')
    
    # Build prerequisite graph data for visualization
    # Format: { course_id: { code, title, prerequisites: [course_ids] } }
    courses_in_program = CurriculumCourse.objects.filter(
        program=program
    ).select_related('course').values_list('course__id', 'course__code', 'course__title')
    
    graph_data = {}
    for course_id, code, title in courses_in_program:
        graph_data[course_id] = {
            'id': course_id,
            'code': code,
            'title': title,
            'prerequisites': []
        }
    
    # Add prerequisite relationships
    for prereq in prerequisites:
        if prereq.course.id in graph_data:
            graph_data[prereq.course.id]['prerequisites'].append(prereq.prerequisite_course.id)
    
    import json
    graph_json = json.dumps(list(graph_data.values()))
    
    ctx = {
        'program': program,
        'prerequisites': prerequisites,
        'graph_json': graph_json,
        'courses': Course.objects.all().order_by('code'),
    }
    return render(request, 'academics/prerequisites_manage.html', ctx)


@admin_required
def prerequisite_add(request, program_id):
    """Add a prerequisite relationship"""
    program = get_object_or_404(Program, id=program_id)
    
    if request.method == 'POST':
        try:
            course_id = request.POST.get('course_id')
            prereq_course_id = request.POST.get('prerequisite_course_id')
            
            course = Course.objects.get(id=course_id)
            prereq_course = Course.objects.get(id=prereq_course_id)
            
            if course.id == prereq_course.id:
                messages.error(request, 'A course cannot be its own prerequisite.')
                return redirect('academics:prerequisites_manage', program_id=program.id)
            
            prerequisite, created = Prerequisite.objects.get_or_create(
                program=program,
                course=course,
                prerequisite_course=prereq_course
            )
            
            if created:
                messages.success(request, f'{course.code} now requires {prereq_course.code}.')
            else:
                messages.info(request, 'This prerequisite already exists.')
            
            return redirect('academics:prerequisites_manage', program_id=program.id)
            
        except Course.DoesNotExist:
            messages.error(request, 'Invalid course selected.')
            return redirect('academics:prerequisites_manage', program_id=program.id)
    
    return redirect('academics:prerequisites_manage', program_id=program.id)


@admin_required
def prerequisite_remove(request, program_id, prereq_id):
    """Remove a prerequisite relationship"""
    program = get_object_or_404(Program, id=program_id)
    prerequisite = get_object_or_404(Prerequisite, id=prereq_id, program=program)
    
    if request.method == 'POST':
        course_code = prerequisite.course.code
        prereq_code = prerequisite.prerequisite_course.code
        prerequisite.delete()
        messages.success(request, f'Removed prerequisite: {course_code} no longer requires {prereq_code}.')
        return redirect('academics:prerequisites_manage', program_id=program.id)
    
    return redirect('academics:prerequisites_manage', program_id=program.id)


@admin_required
def program_import(request):
    """Bulk program import from CSV/Excel/JSON"""
    import csv
    import json
    from io import TextIOWrapper
    
    if request.method == 'POST':
        uploaded_file = request.FILES.get('import_file')
        
        if not uploaded_file:
            messages.error(request, 'Please select a file to import.')
            return redirect('academics:program_import')
        
        file_ext = uploaded_file.name.split('.')[-1].lower()
        
        try:
            imported = 0
            errors = []
            
            if file_ext == 'csv':
                file_data = TextIOWrapper(uploaded_file.file, encoding='utf-8')
                reader = csv.DictReader(file_data)
                
                for idx, row in enumerate(reader, start=2):
                    result = _import_program_row(row, idx)
                    if result['success']:
                        imported += 1
                    else:
                        errors.append(result['error'])
                        
            elif file_ext in ['xlsx', 'xls']:
                from openpyxl import load_workbook
                workbook = load_workbook(uploaded_file)
                sheet = workbook.active
                header = [cell.value for cell in sheet[1]]
                
                for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    row_data = {header[i]: cell.value for i, cell in enumerate(row) if i < len(header)}
                    result = _import_program_row(row_data, idx)
                    if result['success']:
                        imported += 1
                    else:
                        errors.append(result['error'])
                        
            elif file_ext == 'json':
                data = json.load(uploaded_file)
                for idx, row in enumerate(data, start=1):
                    result = _import_program_row(row, idx)
                    if result['success']:
                        imported += 1
                    else:
                        errors.append(result['error'])
            else:
                messages.error(request, 'Unsupported file format.')
                return redirect('academics:program_import')
            
            if imported > 0:
                messages.success(request, f'Successfully imported {imported} program(s).')
            
            if errors:
                messages.warning(request, f'{len(errors)} error(s) occurred.')
            
            return redirect('academics:program_list')
            
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('academics:program_import')
    
    return render(request, 'academics/program_import.html')


@admin_required
def copy_curriculum(request, program_id):
    """Copy curriculum from another program"""
    target_program = get_object_or_404(Program, id=program_id)
    
    if request.method == 'POST':
        source_program_id = request.POST.get('source_program_id')
        overwrite = request.POST.get('overwrite') == 'on'
        
        try:
            source_program = Program.objects.get(id=source_program_id)
            
            # Get source curriculum
            source_curriculum = CurriculumCourse.objects.filter(
                program=source_program
            ).select_related('course', 'level')
            
            if not source_curriculum.exists():
                messages.warning(request, f'{source_program.code} has no curriculum to copy.')
                return redirect('academics:program_detail', program_id=target_program.id)
            
            # Clear target curriculum if overwrite
            if overwrite:
                deleted = CurriculumCourse.objects.filter(program=target_program).delete()[0]
                messages.info(request, f'Removed {deleted} existing course(s) from {target_program.code}.')
            
            # Copy courses
            copied = 0
            skipped = 0
            
            for entry in source_curriculum:
                _, created = CurriculumCourse.objects.get_or_create(
                    program=target_program,
                    course=entry.course,
                    level=entry.level,
                    semester=entry.semester,
                    defaults={'is_compulsory': entry.is_compulsory}
                )
                if created:
                    copied += 1
                else:
                    skipped += 1
            
            messages.success(request, f'Copied {copied} course(s) from {source_program.code}. {skipped} already existed.')
            return redirect('academics:program_detail', program_id=target_program.id)
            
        except Program.DoesNotExist:
            messages.error(request, 'Invalid source program selected.')
            return redirect('academics:program_detail', program_id=target_program.id)
    
    # GET - show form
    programs = Program.objects.exclude(id=target_program.id).order_by('name')
    
    ctx = {
        'target_program': target_program,
        'programs': programs,
    }
    return render(request, 'academics/copy_curriculum.html', ctx)


def _import_program_row(row, row_num):
    """Helper to import a single program"""
    try:
        code = row.get('code')
        name = row.get('name')
        
        if not code or not name:
            return {'success': False, 'error': f'Row {row_num}: Missing required fields'}
        
        min_units = int(row.get('min_units_to_graduate', 120))
        scheme = row.get('classification_scheme', Program.CLASS_SCHEME_BSC)
        
        if scheme not in [Program.CLASS_SCHEME_BSC, Program.CLASS_SCHEME_ND]:
            scheme = Program.CLASS_SCHEME_BSC
        
        program, created = Program.objects.update_or_create(
            code=code,
            defaults={
                'name': name,
                'min_units_to_graduate': min_units,
                'classification_scheme': scheme,
            }
        )
        
        return {'success': True, 'program': program}
        
    except Exception as e:
        return {'success': False, 'error': f'Row {row_num}: {str(e)}'}
