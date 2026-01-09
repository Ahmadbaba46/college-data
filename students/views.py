from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q

from students.models import Student, Level, Session
from academics.models import Program
from users.decorators import admin_or_data_entry_required, role_required


@login_required
def student_list(request):
    """
    List all students with search and filters.
    All authenticated users can view (but only Admin/DataEntry can edit).
    """
    students = Student.objects.select_related(
        'program', 'current_level', 'current_session'
    ).order_by('-created_at')
    
    # Search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Filters
    program_id = request.GET.get('program')
    if program_id:
        students = students.filter(program_id=program_id)
    
    level_id = request.GET.get('level')
    if level_id:
        students = students.filter(current_level_id=level_id)
    
    session_id = request.GET.get('session')
    if session_id:
        students = students.filter(current_session_id=session_id)
    
    # Pagination
    paginator = Paginator(students, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'programs': Program.objects.all(),
        'levels': Level.objects.all(),
        'sessions': Session.objects.all(),
        'search_query': search_query,
        'total_count': students.count(),
    }
    
    return render(request, 'students/list.html', context)


@login_required
def student_search_partial(request):
    """HTMX endpoint for live search"""
    search_query = request.GET.get('search', '').strip()
    
    students = Student.objects.select_related(
        'program', 'current_level', 'current_session'
    )
    
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    students = students.order_by('-created_at')[:10]
    
    return render(request, 'students/partials/student_rows.html', {'students': students})


@login_required
def student_detail(request, student_id):
    """
    View student details with tabs.
    All authenticated users can view.
    """
    student = get_object_or_404(
        Student.objects.select_related('program', 'current_level', 'current_session'),
        id=student_id
    )
    
    # Calculate academic metrics
    from academics.graduation import compute_cgpa_for_program, classify
    from configuration.models import AcademicPolicySettings
    from grading.models import Enrollment, Grade
    from django.db.models import Q
    
    policy = AcademicPolicySettings.get_solo()
    cgpa, total_units = compute_cgpa_for_program(student, require_approved=policy.require_approved_for_metrics)
    
    classification = None
    if student.program:
        classification = classify(student.program, cgpa)
    
    # Get current enrollments
    current_enrollments = Enrollment.objects.filter(
        student=student
    ).select_related(
        'course_offering__course',
        'course_offering__session'
    ).order_by('-course_offering__session__name')[:10]
    
    # Get all grades grouped by session/semester
    grades = Grade.objects.filter(
        enrollment__student=student,
        status=Grade.STATUS_APPROVED
    ).select_related(
        'enrollment__course_offering__course',
        'enrollment__course_offering__session'
    ).order_by('-enrollment__course_offering__session__name', 'enrollment__course_offering__course__code')
    
    # Group grades by session
    grades_by_session = {}
    for grade in grades:
        session_name = grade.enrollment.course_offering.session.name if grade.enrollment.course_offering.session else 'Unknown'
        if session_name not in grades_by_session:
            grades_by_session[session_name] = []
        grades_by_session[session_name].append(grade)
    
    context = {
        'student': student,
        'cgpa': cgpa,
        'total_units': total_units,
        'classification': classification,
        'current_enrollments': current_enrollments,
        'grades_by_session': grades_by_session,
        'total_enrollments': current_enrollments.count(),
    }
    
    return render(request, 'students/detail.html', context)


@admin_or_data_entry_required
def student_create(request):
    """
    Create new student.
    Only Admin and DataEntry can access.
    """
    if request.method == 'POST':
        try:
            student = Student.objects.create(
                student_id=request.POST.get('student_id'),
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                email=request.POST.get('email'),
                phone=request.POST.get('phone', ''),
                date_of_birth=request.POST.get('date_of_birth') or None,
                gender=request.POST.get('gender', ''),
                address=request.POST.get('address', ''),
                program_id=request.POST.get('program') or None,
                current_level_id=request.POST.get('current_level') or None,
                current_session_id=request.POST.get('current_session') or None,
                entry_session_id=request.POST.get('entry_session') or None,
            )
            
            # Handle photo upload
            if 'photo' in request.FILES:
                student.photo = request.FILES['photo']
                student.save()
            
            messages.success(request, f'Student {student.full_name} created successfully!')
            return redirect('students:detail', student_id=student.id)
        
        except Exception as e:
            messages.error(request, f'Error creating student: {str(e)}')
    
    context = {
        'programs': Program.objects.all(),
        'levels': Level.objects.all(),
        'sessions': Session.objects.all(),
    }
    
    return render(request, 'students/create.html', context)


@admin_or_data_entry_required
def student_edit(request, student_id):
    """
    Edit existing student.
    Only Admin and DataEntry can access.
    """
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        try:
            student.student_id = request.POST.get('student_id')
            student.first_name = request.POST.get('first_name')
            student.last_name = request.POST.get('last_name')
            student.email = request.POST.get('email')
            student.phone = request.POST.get('phone', '')
            student.date_of_birth = request.POST.get('date_of_birth') or None
            student.gender = request.POST.get('gender', '')
            student.address = request.POST.get('address', '')
            student.program_id = request.POST.get('program') or None
            student.current_level_id = request.POST.get('current_level') or None
            student.current_session_id = request.POST.get('current_session') or None
            
            # Handle photo upload
            if 'photo' in request.FILES:
                student.photo = request.FILES['photo']
            
            student.save()
            
            messages.success(request, f'Student {student.full_name} updated successfully!')
            return redirect('students:detail', student_id=student.id)
        
        except Exception as e:
            messages.error(request, f'Error updating student: {str(e)}')
    
    context = {
        'student': student,
        'programs': Program.objects.all(),
        'levels': Level.objects.all(),
        'sessions': Session.objects.all(),
    }
    
    return render(request, 'students/edit.html', context)


@admin_or_data_entry_required
def student_delete(request, student_id):
    """
    Delete student.
    Only Admin and DataEntry can access.
    """
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        student_name = student.full_name
        student.delete()
        messages.success(request, f'Student {student_name} deleted successfully!')
        return redirect('students:list')
    
    return render(request, 'students/delete_confirm.html', {'student': student})


@role_required('Admin', 'DataEntry')
def student_promote(request):
    """Bulk student promotion page"""
    from students.models import Level, Session
    from django.contrib import messages
    
    if request.method == 'POST':
        current_level_id = request.POST.get('current_level')
        current_session_id = request.POST.get('current_session')
        next_level_id = request.POST.get('next_level')
        next_session_id = request.POST.get('next_session')
        student_ids = request.POST.getlist('student_ids')
        
        if not all([current_level_id, current_session_id, next_level_id, next_session_id]):
            messages.error(request, 'Please fill all required fields.')
            return redirect('students:promote')
        
        try:
            current_level = Level.objects.get(id=current_level_id)
            current_session = Session.objects.get(id=current_session_id)
            next_level = Level.objects.get(id=next_level_id)
            next_session = Session.objects.get(id=next_session_id)
            
            # Get students to promote
            students_to_promote = Student.objects.filter(
                current_level=current_level,
                current_session=current_session
            )
            
            # If specific students selected, filter by them
            if student_ids:
                students_to_promote = students_to_promote.filter(id__in=student_ids)
            
            if not students_to_promote.exists():
                messages.warning(request, 'No students found to promote with the given criteria.')
                return redirect('students:promote')
            
            # Perform promotion
            promoted_count = students_to_promote.update(
                current_level=next_level,
                current_session=next_session
            )
            
            messages.success(request, f'Successfully promoted {promoted_count} student(s) from {current_level.name} to {next_level.name}.')
            return redirect('students:list')
            
        except (Level.DoesNotExist, Session.DoesNotExist) as e:
            messages.error(request, f'Invalid level or session selected.')
            return redirect('students:promote')
    
    # GET request - show form
    levels = Level.objects.all().order_by('name')
    sessions = Session.objects.all().order_by('-name')
    
    ctx = {
        'levels': levels,
        'sessions': sessions,
    }
    return render(request, 'students/promote.html', ctx)


@login_required
def promotion_preview_partial(request):
    """HTMX partial to preview students for promotion"""
    current_level_id = request.GET.get('current_level')
    current_session_id = request.GET.get('current_session')
    
    if not current_level_id or not current_session_id:
        return render(request, 'students/partials/promotion_preview.html', {'students': []})
    
    try:
        current_level = Level.objects.get(id=current_level_id)
        current_session = Session.objects.get(id=current_session_id)
        
        students = Student.objects.filter(
            current_level=current_level,
            current_session=current_session
        ).select_related('program').order_by('student_id')
        
        ctx = {
            'students': students,
            'current_level': current_level,
            'current_session': current_session,
        }
        return render(request, 'students/partials/promotion_preview.html', ctx)
        
    except (Level.DoesNotExist, Session.DoesNotExist):
        return render(request, 'students/partials/promotion_preview.html', {'students': []})


@role_required('Admin', 'DataEntry')
def student_import(request):
    """Bulk student import from CSV/Excel/JSON"""
    import csv
    import json
    from io import TextIOWrapper
    
    if request.method == 'POST':
        uploaded_file = request.FILES.get('import_file')
        
        if not uploaded_file:
            messages.error(request, 'Please select a file to import.')
            return redirect('students:import')
        
        file_ext = uploaded_file.name.split('.')[-1].lower()
        
        # Parse file based on extension
        try:
            imported = 0
            errors = []
            
            if file_ext == 'csv':
                # Handle CSV
                file_data = TextIOWrapper(uploaded_file.file, encoding='utf-8')
                reader = csv.DictReader(file_data)
                
                for idx, row in enumerate(reader, start=2):
                    result = _import_student_row(row, idx)
                    if result['success']:
                        imported += 1
                    else:
                        errors.append(result['error'])
                        
            elif file_ext in ['xlsx', 'xls']:
                # Handle Excel
                from openpyxl import load_workbook
                workbook = load_workbook(uploaded_file)
                sheet = workbook.active
                header = [cell.value for cell in sheet[1]]
                
                for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    row_data = {header[i]: cell.value for i, cell in enumerate(row) if i < len(header)}
                    result = _import_student_row(row_data, idx)
                    if result['success']:
                        imported += 1
                    else:
                        errors.append(result['error'])
                        
            elif file_ext == 'json':
                # Handle JSON
                data = json.load(uploaded_file)
                for idx, row in enumerate(data, start=1):
                    result = _import_student_row(row, idx)
                    if result['success']:
                        imported += 1
                    else:
                        errors.append(result['error'])
            else:
                messages.error(request, 'Unsupported file format. Please use CSV, Excel, or JSON.')
                return redirect('students:import')
            
            # Show results
            if imported > 0:
                messages.success(request, f'Successfully imported {imported} student(s).')
            
            if errors:
                error_msg = f'{len(errors)} error(s) occurred during import. '
                messages.warning(request, error_msg + ' | '.join(errors[:5]))
                if len(errors) > 5:
                    messages.info(request, f'... and {len(errors) - 5} more errors.')
            
            return redirect('students:list')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('students:import')
    
    # GET request - show form with sample templates
    return render(request, 'students/import.html')


@admin_or_data_entry_required
def bulk_assign_program(request):
    """Bulk assign students to a program"""
    from academics.models import Program
    
    if request.method == 'POST':
        program_id = request.POST.get('program_id')
        student_ids = request.POST.getlist('student_ids')
        
        if not program_id:
            messages.error(request, 'Please select a program.')
            return redirect('students:bulk_assign_program')
        
        if not student_ids:
            messages.error(request, 'Please select at least one student.')
            return redirect('students:bulk_assign_program')
        
        try:
            program = Program.objects.get(id=program_id)
            updated = Student.objects.filter(id__in=student_ids).update(program=program)
            messages.success(request, f'Assigned {updated} student(s) to {program.code}.')
            return redirect('students:list')
        except Program.DoesNotExist:
            messages.error(request, 'Invalid program selected.')
            return redirect('students:bulk_assign_program')
    
    # GET - show form
    from academics.models import Program
    programs = Program.objects.all().order_by('name')
    
    # Get students with optional filters
    students = Student.objects.all().order_by('student_id')
    
    # Filter by level
    level_id = request.GET.get('level')
    if level_id:
        students = students.filter(current_level_id=level_id)
    
    # Filter by current program
    filter_program = request.GET.get('filter_program')
    if filter_program:
        students = students.filter(program_id=filter_program)
    elif filter_program == '':
        # Show only unassigned
        students = students.filter(program__isnull=True)
    
    # Search
    search = request.GET.get('search', '').strip()
    if search:
        students = students.filter(
            Q(student_id__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )
    
    context = {
        'students': students[:100],  # Limit for performance
        'programs': programs,
        'levels': Level.objects.all(),
        'total_count': students.count(),
    }
    
    return render(request, 'students/bulk_assign_program.html', context)


def _import_student_row(row, row_num):
    """Helper function to import a single student row"""
    try:
        # Required fields
        required_fields = ['student_id', 'first_name', 'last_name']
        for field in required_fields:
            if not row.get(field):
                return {'success': False, 'error': f'Row {row_num}: Missing required field "{field}"'}
        
        # Get or create related objects
        entry_level = None
        if row.get('entry_level_id'):
            try:
                entry_level = Level.objects.get(id=row['entry_level_id'])
            except Level.DoesNotExist:
                return {'success': False, 'error': f'Row {row_num}: Invalid entry_level_id'}
        
        current_level = None
        if row.get('current_level_id'):
            try:
                current_level = Level.objects.get(id=row['current_level_id'])
            except Level.DoesNotExist:
                return {'success': False, 'error': f'Row {row_num}: Invalid current_level_id'}
        
        current_session = None
        if row.get('current_session_id'):
            try:
                current_session = Session.objects.get(id=row['current_session_id'])
            except Session.DoesNotExist:
                return {'success': False, 'error': f'Row {row_num}: Invalid current_session_id'}
        
        program = None
        if row.get('program_id'):
            try:
                program = Program.objects.get(id=row['program_id'])
            except Program.DoesNotExist:
                return {'success': False, 'error': f'Row {row_num}: Invalid program_id'}
        
        # Create or update student
        student, created = Student.objects.update_or_create(
            student_id=row['student_id'],
            defaults={
                'first_name': row['first_name'],
                'last_name': row['last_name'],
                'email': row.get('email', ''),
                'phone': row.get('phone', ''),
                'entry_level': entry_level,
                'current_level': current_level,
                'current_session': current_session,
                'program': program,
            }
        )
        
        return {'success': True, 'student': student, 'created': created}
        
    except Exception as e:
        return {'success': False, 'error': f'Row {row_num}: {str(e)}'}
