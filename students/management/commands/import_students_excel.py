from django.core.management.base import BaseCommand
from students.models import Student, Level, Session
from openpyxl import load_workbook

class Command(BaseCommand):
    help = 'Import students from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='The path to the Excel file')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']

        try:
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active

            header = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2):
                row_data = {header[i]: cell.value for i, cell in enumerate(row)}

                try:
                    entry_level = Level.objects.get(id=row_data['entry_level_id'])
                    current_level = Level.objects.get(id=row_data['current_level_id'])
                    current_session = Session.objects.get(id=row_data['current_session_id'])

                    student, created = Student.objects.get_or_create(
                        student_id=row_data['student_id'],
                        defaults={
                            'first_name': row_data['first_name'],
                            'last_name': row_data['last_name'],
                            'entry_level': entry_level,
                            'current_level': current_level,
                            'current_session': current_session
                        }
                    )

                    if created:
                        self.stdout.write(self.style.SUCCESS(f'Successfully created student: {student}'))
                    else:
                        self.stdout.write(self.style.WARNING(f'Student with ID {row_data["student_id"]} already exists.'))

                except Level.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'Invalid Level ID in row: {row_data}'))
                except Session.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'Invalid Session ID in row: {row_data}'))

        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'File not found at: {file_path}'))
